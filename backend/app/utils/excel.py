"""Excel export utility"""
from typing import List, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO


def export_cases_to_excel(cases: List[Dict[str, Any]], fields: List[Dict[str, Any]]) -> bytes:
    """
    Export test cases to Excel file
    
    Args:
        cases: List of test case data
        fields: List of field configurations
    
    Returns:
        Excel file bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Get field labels
    field_labels = {f["field_name"]: f["field_label"] for f in fields}
    visible_fields = [f for f in fields if f.get("is_visible", True)]
    
    # Write headers
    headers = ["序号"] + [f["field_label"] for f in visible_fields]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row_idx, case in enumerate(cases, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        
        for col_idx, field in enumerate(visible_fields, 2):
            field_name = field["field_name"]
            value = case.get("case_data", {}).get(field_name, "")
            
            # Handle different types
            if isinstance(value, (dict, list)):
                value = str(value)
            
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def export_case_version_comparison(case_v1: Dict[str, Any], case_v2: Dict[str, Any], 
                                    fields: List[Dict[str, Any]]) -> bytes:
    """
    Export version comparison to Excel
    
    Args:
        case_v1: Previous version data
        case_v2: Current version data
        fields: Field configurations
    
    Returns:
        Excel file bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "版本对比"
    
    # Header style
    header_font = Font(bold=True)
    
    visible_fields = [f for f in fields if f.get("is_visible", True)]
    
    # Write headers
    ws.cell(row=1, column=1, value="字段").font = header_font
    ws.cell(row=1, column=2, value=f"版本{case_v1.get('version', 'N/A')}").font = header_font
    ws.cell(row=1, column=3, value=f"版本{case_v2.get('version', 'N/A')}").font = header_font
    
    # Write data
    for row_idx, field in enumerate(visible_fields, 2):
        field_name = field["field_name"]
        field_label = field["field_label"]
        
        value1 = case_v1.get("case_data", {}).get(field_name, "")
        value2 = case_v2.get("case_data", {}).get(field_name, "")
        
        ws.cell(row=row_idx, column=1, value=field_label)
        ws.cell(row=row_idx, column=2, value=str(value1))
        ws.cell(row=row_idx, column=3, value=str(value2))
        
        # Highlight changes
        if value1 != value2:
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
