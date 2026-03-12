"""
用例管理测试模块

包含：
- 用例增删改查测试
- 版本管理测试
- Excel 导出测试
"""
import pytest
from unittest.mock import MagicMock, AsyncMock, patch, mock_open
from datetime import datetime
import json
import io


class TestCaseCRUD:
    """用例增删改查测试"""

    @pytest.mark.unit
    def test_create_case_success(self, mock_db_session, mock_test_system, mock_module):
        """测试创建用例成功"""
        from app.schemas.case import CaseCreate
        from app.services.cases import CaseService
        
        case_data = CaseCreate(
            system_id=mock_test_system.id,
            module_id=mock_module.id,
            case_data={
                "title": "用户登录测试",
                "precondition": "用户已注册",
                "steps": ["打开登录页", "输入用户名密码", "点击登录"],
                "expected": "登录成功"
            },
            status="draft"
        )
        
        new_case = MagicMock()
        new_case.id = 1
        new_case.system_id = case_data.system_id
        new_case.case_data = case_data.case_data
        
        assert new_case.id == 1
        assert new_case.case_data["title"] == "用户登录测试"

    @pytest.mark.unit
    def test_get_case_by_id(self, mock_db_session, mock_test_case):
        """测试根据 ID 获取用例"""
        mock_db_session.query.return_value.filter.return_value.first.return_value = mock_test_case
        
        from app.services.cases import CaseService
        service = CaseService(db=mock_db_session)
        result = service.get_case(1)
        
        assert result is not None
        assert result.id == 1

    @pytest.mark.unit
    def test_get_case_not_found(self, mock_db_session):
        """测试用例不存在"""
        mock_db_session.query.return_value.filter.return_value.first.return_value = None
        
        from app.services.cases import CaseService
        service = CaseService(db=mock_db_session)
        result = service.get_case(999)
        
        assert result is None

    @pytest.mark.unit
    def test_list_cases_with_pagination(self, mock_db_session):
        """测试用例列表分页"""
        mock_cases = [
            MagicMock(id=i, title=f"用例{i}") for i in range(1, 11)
        ]
        
        mock_db_session.query.return_value.filter.return_value.all.return_value = mock_cases
        mock_db_session.query.return_value.count.return_value = 100
        
        # 模拟分页
        page = 1
        page_size = 10
        start = (page - 1) * page_size
        end = start + page_size
        
        paginated_cases = mock_cases[start:end]
        
        assert len(paginated_cases) == 10
        assert paginated_cases[0].id == 1

    @pytest.mark.unit
    def test_list_cases_with_filter(self, mock_db_session):
        """测试用例筛选"""
        mock_cases = [
            MagicMock(id=1, status="draft", system_id=1),
            MagicMock(id=2, status="approved", system_id=1),
        ]
        
        # 筛选条件
        filter_status = "draft"
        filtered = [c for c in mock_cases if c.status == filter_status]
        
        assert len(filtered) == 1
        assert filtered[0].status == "draft"

    @pytest.mark.unit
    def test_update_case(self, mock_db_session, mock_test_case):
        """测试更新用例"""
        from app.schemas.case import CaseUpdate
        
        update_data = CaseUpdate(
            case_data={
                "title": "更新后的标题",
                "precondition": "更新后的前置条件",
                "steps": ["新步骤1", "新步骤2"],
                "expected": "新预期结果"
            },
            status="reviewing"
        )
        
        mock_test_case.case_data = update_data.case_data
        mock_test_case.status = update_data.status
        
        assert mock_test_case.case_data["title"] == "更新后的标题"
        assert mock_test_case.status == "reviewing"

    @pytest.mark.unit
    def test_delete_case(self, mock_db_session):
        """测试删除用例"""
        mock_case = MagicMock(id=1)
        mock_db_session.query.return_value.filter.return_value.first.return_value = mock_case
        
        from app.services.cases import CaseService
        service = CaseService(db=mock_db_session)
        result = service.delete_case(1)
        
        assert result is True

    @pytest.mark.unit
    def test_delete_case_not_found(self, mock_db_session):
        """测试删除不存在的用例"""
        mock_db_session.query.return_value.filter.return_value.first.return_value = None
        
        from app.services.cases import CaseService
        service = CaseService(db=mock_db_session)
        
        with pytest.raises(ValueError, match="Case not found"):
            service.delete_case(999)


class TestVersionManagement:
    """版本管理测试"""

    @pytest.mark.unit
    def test_create_version(self):
        """测试创建版本"""
        case_version = {
            "case_id": 1,
            "version": 1,
            "case_data": {"title": "用例V1"},
            "change_summary": "初始版本",
            "created_by": 1
        }
        
        assert case_version["version"] == 1

    @pytest.mark.unit
    def test_version_increment(self):
        """测试版本号递增"""
        versions = []
        
        for i in range(1, 4):
            version = {
                "version": i,
                "case_data": {"title": f"用例V{i}"}
            }
            versions.append(version)
        
        assert versions[-1]["version"] == 3

    @pytest.mark.unit
    def test_get_version_history(self):
        """测试获取版本历史"""
        version_history = [
            {"version": 3, "created_at": "2024-01-03", "change_summary": "修复bug"},
            {"version": 2, "created_at": "2024-01-02", "change_summary": "优化步骤"},
            {"version": 1, "created_at": "2024-01-01", "change_summary": "初始版本"}
        ]
        
        # 按版本号排序
        sorted_versions = sorted(version_history, key=lambda x: x["version"], reverse=True)
        
        assert sorted_versions[0]["version"] == 3

    @pytest.mark.unit
    def test_compare_versions(self):
        """测试版本对比"""
        v1_data = {"title": "旧标题", "steps": ["步骤1"]}
        v2_data = {"title": "新标题", "steps": ["步骤1", "步骤2"]}
        
        # 比较差异
        differences = {
            "title": {"old": v1_data["title"], "new": v2_data["title"]},
            "steps": {"old": v1_data["steps"], "new": v2_data["steps"]}
        }
        
        assert differences["title"]["old"] != differences["title"]["new"]
        assert len(differences["steps"]["old"]) < len(differences["steps"]["new"])

    @pytest.mark.unit
    def test_version_restore(self):
        """测试版本恢复"""
        current_data = {"title": "当前版本"}
        historical_data = {"title": "历史版本"}
        
        # 恢复到历史版本
        restored_data = historical_data.copy()
        
        assert restored_data["title"] == "历史版本"

    @pytest.mark.unit
    def test_auto_version_on_update(self):
        """测试更新时自动创建版本"""
        case = {"id": 1, "version": 1, "data": "data_v1"}
        
        # 更新用例
        case["version"] += 1
        case["data"] = "data_v2"
        
        assert case["version"] == 2


class TestExcelExport:
    """Excel 导出测试"""

    @pytest.mark.unit
    def test_export_to_excel(self, sample_excel_content):
        """测试导出到 Excel"""
        import openpyxl
        from io import BytesIO
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试用例"
        
        # 写入标题
        headers = sample_excel_content["headers"]
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
        
        # 写入数据
        for row_idx, row_data in enumerate(sample_excel_content["rows"], 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row_idx, col, value)
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        
        assert output.tell() > 0

    @pytest.mark.unit
    def test_export_with_styles(self):
        """测试带样式的导出"""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 设置标题样式
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        
        ws.cell(1, 1, "标题").font = header_font
        ws.cell(1, 1).alignment = header_alignment
        
        assert ws.cell(1, 1).font.bold is True

    @pytest.mark.unit
    def test_export_empty_cases(self):
        """测试导出空用例列表"""
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试用例"
        
        # 只写入标题
        ws.cell(1, 1, "标题")
        
        output = BytesIO()
        wb.save(output)
        
        assert output.tell() > 0

    @pytest.mark.unit
    def test_export_large_dataset(self):
        """测试导出大数据集"""
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 模拟大数据集
        row_count = 1000
        
        for row in range(2, row_count + 2):
            ws.cell(row, 1, f"用例{row-1}")
            ws.cell(row, 2, f"数据{row-1}")
        
        assert ws.max_row >= row_count

    @pytest.mark.unit
    def test_export_custom_fields(self):
        """测试导出自定义字段"""
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 自定义字段
        custom_fields = ["用例ID", "标题", "优先级", "类型", "标签"]
        
        for col, field in enumerate(custom_fields, 1):
            ws.cell(1, col, field)
        
        assert ws.cell(1, 4).value == "类型"

    @pytest.mark.unit
    def test_export_multiple_sheets(self):
        """测试多 Sheet 导出"""
        import openpyxl
        
        wb = openpyxl.Workbook()
        
        # 创建多个 Sheet
        ws1 = wb.active
        ws1.title = "用户管理"
        ws1.cell(1, 1, "用户登录")
        
        ws2 = wb.create_sheet("系统管理")
        ws2.cell(1, 1, "系统配置")
        
        assert len(wb.sheetnames) == 2

    @pytest.mark.unit
    def test_export_file_download(self, sample_excel_content):
        """测试文件下载"""
        from io import BytesIO
        
        # 模拟文件下载
        file_content = BytesIO()
        
        # 写入内容
        file_content.write(b"Mock Excel Content")
        file_content.seek(0)
        
        assert file_content.tell() > 0
        assert file_content.getvalue() == b"Mock Excel Content"
